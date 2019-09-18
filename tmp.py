# ！/usr/local/env python
# encoding=utf8
import openpyxl
import telnetlib

def write_info(host, port,flag,server,comm,comm_version):
    with open("check_info.csv",'a+') as f:
        f.writelines("{},{},{},{},{}\n".format(host,port,flag,server,comm,comm_version))

def conn_scan(host, port,server,comm,comm_version):
    t = telnetlib.Telnet()
    try:
        t.open(host, port, timeout=10)
        flag = "验证存在"
    except   Exception:
        # print(host, port, 'is not avaliable')
        flag = "验证不存在"
    finally:
        write_info(host, port, flag, server, comm, comm_version)
        t.close()


def tmp():
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook('tmp.xlsx')
    # sheets = wb.sheetnames
    # print(sheets, type(sheets))
    # 获取指定的表单
    # sheet3 = wb['Sheet1']

    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    for i in range(2, 1273):
        port = ws.cell(row=i, column=2).value
        ip = ws.cell(row=i, column=1).value
        server = ws.cell(row=i, column=5).value
        comm = ws.cell(row=i, column=6).value
        comm_version = ws.cell(row=i, column=7).value

        conn_scan(ip,port,server,comm,comm_version)



if __name__ == '__main__':
    tmp()
